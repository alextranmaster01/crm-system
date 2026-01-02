import streamlit as st
import pandas as pd
import datetime
from datetime import datetime, timedelta
import re
import io
import time
import json
import mimetypes
import numpy as np

# =============================================================================
# 1. C·∫§U H√åNH & KH·ªûI T·∫†O
# =============================================================================
APP_VERSION = "V6025 - FIX QUOTE TOTAL & EXPORT ROW"
st.set_page_config(page_title=f"CRM {APP_VERSION}", layout="wide", page_icon="üíé")

# CSS UI
st.markdown("""
    <style>
    button[data-baseweb="tab"] div p { font-size: 18px !important; font-weight: 700 !important; }
    .card-3d { border-radius: 12px; padding: 20px; color: white; text-align: center; box-shadow: 0 4px 8px rgba(0,0,0,0.2); margin-bottom: 10px; }
    .bg-sales { background: linear-gradient(135deg, #00b09b, #96c93d); }
    .bg-cost { background: linear-gradient(135deg, #ff5f6d, #ffc371); }
    .bg-profit { background: linear-gradient(135deg, #f83600, #f9d423); }
    [data-testid="stDataFrame"] > div { max-height: 750px; }
    .highlight-low { background-color: #ffcccc !important; color: red !important; font-weight: bold; }
    
    /* CSS CHO C√ÅC N√öT B·∫§M: N·ªÄN T·ªêI - CH·ªÆ S√ÅNG */
    div.stButton > button { 
        width: 100%; 
        border-radius: 5px; 
        font-weight: bold; 
        background-color: #262730; /* N·ªÅn t·ªëi */
        color: #ffffff; /* Ch·ªØ tr·∫Øng */
        border: 1px solid #4e4e4e;
    }
    div.stButton > button:hover {
        background-color: #444444;
        color: #ffffff;
        border-color: #ffffff;
    }
    
    /* STYLE CHO TOTAL VIEW */
    .total-view {
        font-size: 20px;
        font-weight: bold;
        color: #00FF00; /* M√†u xanh l√° n·ªïi b·∫≠t */
        background-color: #262730;
        padding: 10px;
        border-radius: 8px;
        text-align: right;
        margin-top: 10px;
        border: 1px solid #4e4e4e;
    }

    /* --- FIX: STYLE CHO D√íNG TOTAL (D√íNG CU·ªêI C√ôNG TRONG TABLE) M√ÄU V√ÄNG --- */
    /* L∆∞u √Ω: ƒê√¢y l√† CSS hack ƒë·ªÉ highlight d√≤ng cu·ªëi c·ªßa b·∫£ng hi·ªÉn th·ªã */
    /* √Åp d·ª•ng cho c·∫£ st.dataframe v√† st.data_editor */
    [data-testid="stDataFrame"] table tbody tr:last-child {
        background-color: #FFD700 !important; /* M√†u v√†ng */
        color: #000000 !important; /* Ch·ªØ ƒëen */
        font-weight: 900 !important;
    }
    [data-testid="stDataFrame"] table tbody tr:last-child td {
        color: #000000 !important;
        background-color: #FFD700 !important; /* Force n·ªÅn v√†ng cho t·ª´ng √¥ */
        font-weight: bold !important;
    }
    </style>""", unsafe_allow_html=True)

# LIBRARIES & CONNECTIONS
try:
    from supabase import create_client, Client
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
except ImportError:
    st.error("‚ö†Ô∏è Thi·∫øu th∆∞ vi·ªán. Vui l√≤ng ch·∫°y l·ªánh: pip install streamlit pandas supabase google-api-python-client google-auth-oauthlib openpyxl")
    st.stop()

# CONNECT SERVER
try:
    if "supabase" not in st.secrets or "google_oauth" not in st.secrets:
        st.error("‚ö†Ô∏è Ch∆∞a c·∫•u h√¨nh secrets.toml. Vui l√≤ng ki·ªÉm tra l·∫°i file secrets.")
        st.stop()

    SUPABASE_URL = st.secrets["supabase"]["url"]
    SUPABASE_KEY = st.secrets["supabase"]["key"]
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    OAUTH_INFO = st.secrets["google_oauth"]
    ROOT_FOLDER_ID = OAUTH_INFO.get("root_folder_id", "1GLhnSK7Bz7LbTC-Q7aPt_Itmutni5Rqa")
except Exception as e:
    st.error(f"‚ö†Ô∏è L·ªói Config: {e}"); st.stop()

# =============================================================================
# 2. H√ÄM H·ªñ TR·ª¢ (UTILS)
# =============================================================================

def get_drive_service():
    try:
        creds = Credentials(None, refresh_token=OAUTH_INFO["refresh_token"], 
                            token_uri="https://oauth2.googleapis.com/token", 
                            client_id=OAUTH_INFO["client_id"], client_secret=OAUTH_INFO["client_secret"])
        return build('drive', 'v3', credentials=creds)
    except: return None

# H√†m t·∫°o folder ƒë·ªá quy
def get_or_create_folder_hierarchy(srv, path_list, parent_id):
    current_parent_id = parent_id
    for folder_name in path_list:
        q = f"'{current_parent_id}' in parents and name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        results = srv.files().list(q=q, fields="files(id)").execute().get('files', [])
        
        if results:
            current_parent_id = results[0]['id']
        else:
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                
                'parents': [current_parent_id]
            }
            folder = srv.files().create(body=file_metadata, fields='id').execute()
            current_parent_id = folder.get('id')
            try: srv.permissions().create(fileId=current_parent_id, body={'role': 'reader', 'type': 'anyone'}).execute()
            except: pass
            
    return current_parent_id

def upload_to_drive_structured(file_obj, path_list, file_name):
    srv = get_drive_service()
    if not srv: return "", ""
    try:
        folder_id = get_or_create_folder_hierarchy(srv, path_list, ROOT_FOLDER_ID)
        media = MediaIoBaseUpload(file_obj, mimetype=mimetypes.guess_type(file_name)[0] or 'application/octet-stream', resumable=True)
        file_meta = {'name': file_name, 'parents': [folder_id]}
        q_ex = f"'{folder_id}' in parents and name='{file_name}' and trashed=false"
        exists = srv.files().list(q=q_ex, fields="files(id)").execute().get('files', [])
        if exists:
            file_id = exists[0]['id']
            srv.files().update(fileId=file_id, media_body=media).execute()
        else:
            file_id = srv.files().create(body=file_meta, media_body=media, fields='id').execute()['id']
        try: srv.permissions().create(fileId=file_id, body={'role': 'reader', 'type': 'anyone'}).execute()
        except: pass
        folder_link = f"https://drive.google.com/drive/folders/{folder_id}"
        return folder_link, file_id
    except Exception as e: 
        st.error(f"L·ªói upload Drive: {e}")
        return "", ""

def upload_to_drive_simple(file_obj, sub_folder, file_name):
    srv = get_drive_service()
    if not srv: return "", ""
    try:
        folder_id = get_or_create_folder_hierarchy(srv, [sub_folder], ROOT_FOLDER_ID)
        media = MediaIoBaseUpload(file_obj, mimetype=mimetypes.guess_type(file_name)[0] or 'application/octet-stream', resumable=True)
        file_meta = {'name': file_name, 'parents': [folder_id]}
        q_ex = f"'{folder_id}' in parents and name='{file_name}' and trashed=false"
        exists = srv.files().list(q=q_ex, fields="files(id)").execute().get('files', [])
        if exists:
            file_id = exists[0]['id']
            srv.files().update(fileId=file_id, media_body=media).execute()
        else:
            file_id = srv.files().create(body=file_meta, media_body=media, fields='id').execute()['id']
        try: srv.permissions().create(fileId=file_id, body={'role': 'reader', 'type': 'anyone'}).execute()
        except: pass
        return f"https://drive.google.com/thumbnail?id={file_id}&sz=w200", file_id
    except: return "", ""

def search_file_in_drive_by_name(name_contains):
    srv = get_drive_service()
    if not srv: return None, None, None
    try:
        q = f"name contains '{name_contains}' and trashed=false"
        results = srv.files().list(q=q, fields="files(id, name, parents)").execute().get('files', [])
        if results:
            return results[0]['id'], results[0]['name'], (results[0]['parents'][0] if 'parents' in results[0] else None)
        return None, None, None
    except: return None, None, None

def download_from_drive(file_id):
    srv = get_drive_service()
    if not srv: return None
    try:
        request = srv.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False: status, done = downloader.next_chunk()
        
        # --- FIX QUAN TR·ªåNG: ƒê∆∞a con tr·ªè v·ªÅ ƒë·∫ßu file ƒë·ªÉ pandas ƒë·ªçc ƒë∆∞·ª£c ---
        fh.seek(0) 
        return fh
    except: return None

def safe_str(val):
    if val is None: return ""
    s = str(val).strip()
    if s.lower() in ['nan', 'none', 'null', 'nat', '']: return ""
    return s

def to_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).replace(",", "").replace("¬•", "").replace("$", "").replace("RMB", "").replace("VND", "").replace(" ", "").upper()
    try:
        nums = re.findall(r"[-+]?\d*\.\d+|\d+", s)
        return float(nums[0]) if nums else 0.0
    except: return 0.0

def fmt_num(x): 
    try:
        if x is None: return "0"
        val = float(x)
        if val.is_integer(): return "{:,.0f}".format(val)
        else:
            s = "{:,.3f}".format(val)
            return s.rstrip('0').rstrip('.')
    except: return "0"

# --- NEW: FORMAT 2 DECIMAL PLACES (FOR QUOTE TAB) ---
def fmt_float_2(x):
    try:
        if x is None: return "0.00"
        val = float(x)
        return "{:,.2f}".format(val)
    except: return "0.00"

def clean_key(s): return safe_str(s).lower()

def calc_eta(order_date_str, leadtime_val):
    try:
        if isinstance(order_date_str, datetime): dt_order = order_date_str
        else:
            try: dt_order = datetime.strptime(order_date_str, "%d/%m/%Y")
            except: dt_order = datetime.now()
        lt_str = str(leadtime_val)
        nums = re.findall(r'\d+', lt_str)
        days = int(nums[0]) if nums else 0
        dt_exp = dt_order + timedelta(days=days)
        return dt_exp.strftime("%d/%m/%Y")
    except: return ""

def load_data(table, order_by="id", ascending=True):
    try:
        query = supabase.table(table).select("*")
        
        # --- FIX L·ªñI IMPORT PGRST204: KH√îNG SORT ROW_ORDER N·∫æU C·ªòT KH√îNG T·ªíN T·∫†I ---
        # Thay v√¨ sort c·ª©ng b·∫±ng Supabase, ta s·∫Ω t·∫£i v·ªÅ r·ªìi sort b·∫±ng Pandas cho an to√†n.
        # Ho·∫∑c d√πng try-except ƒë·ªÉ fallback. ·ªû ƒë√¢y ch·ªçn c√°ch ƒë∆°n gi·∫£n: t·∫£i h·∫øt v·ªÅ r·ªìi sort.
        res = query.execute()
        df = pd.DataFrame(res.data)
        
        if not df.empty:
            # Drop c·ªôt 'id' n·∫øu kh√¥ng c·∫ßn thi·∫øt
            if table != "crm_tracking" and 'id' in df.columns: 
                df = df.drop(columns=['id'])
            
            # Sort b·∫±ng Pandas n·∫øu c√≥ c·ªôt order_by
            if order_by in df.columns:
                df = df.sort_values(by=order_by, ascending=ascending)
            
        return df
    except Exception as e:
        # st.error(f"L·ªói load data {table}: {e}") # C√≥ th·ªÉ uncomment ƒë·ªÉ debug
        return pd.DataFrame()

# =============================================================================
# 3. LOGIC T√çNH TO√ÅN CORE (UPDATED: MANUAL OVERRIDE SUPPORT & NEW PROFIT FORMULA)
# =============================================================================
def recalculate_quote_logic(df, params):
    # 1. Chuy·ªÉn ƒë·ªïi d·ªØ li·ªáu sang s·ªë (Float) ƒë·ªÉ t√≠nh to√°n
    cols_money_input = [
        "Q'ty", "Buying price(VND)", "Buying price(RMB)", "Exchange rate",
        "AP price(VND)", "Unit price(VND)", 
        "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
        "Transportation", "Management fee(%)", "Payback(%)"
    ]
    
    # T·∫°o c·ªôt n·∫øu ch∆∞a c√≥ (ƒë·ªÉ tr√°nh l·ªói) v√† chuy·ªÉn sang s·ªë
    for c in cols_money_input:
        if c not in df.columns: df[c] = 0.0
        df[c] = df[c].apply(to_float)

    # 2. T√çNH TO√ÅN C√ÅC C·ªòT TOTAL & LOGIC C∆† B·∫¢N (Lu√¥n ch·∫°y)
    # Buying VND lu√¥n = RMB * Rate 
    df["Buying price(VND)"] = df["Buying price(RMB)"] * df["Exchange rate"]
    
    df["Total buying price(VND)"] = df["Buying price(VND)"] * df["Q'ty"]
    df["Total buying price(rmb)"] = df["Buying price(RMB)"] * df["Q'ty"]
    df["AP total price(VND)"] = df["AP price(VND)"] * df["Q'ty"]
    df["Total price(VND)"] = df["Unit price(VND)"] * df["Q'ty"]
    
    # GAP l√† k·∫øt qu·∫£ t√≠nh to√°n
    df["GAP"] = df["Total price(VND)"] - df["AP total price(VND)"]

    # 3. T√çNH L·ª¢I NHU·∫¨N (PROFIT)
    # --- UPDATED FORMULA ---
    # Profit = Total price - (Total buying price VND + GAP + End user + Buyer + Import tax + VAT + Transportation + Management fee) + Payback
    
    # L∆∞u √Ω: GAP trong c√¥ng th·ª©c n√†y l√† gi√° tr·ªã GAP th√¥ (Total - AP Total) nh∆∞ y√™u c·∫ßu.
    
    # C·ªông d·ªìn c√°c chi ph√≠ (bao g·ªìm GAP)
    cost_ops = (df["Total buying price(VND)"] + 
                df["GAP"] +
                df["End user(%)"] + 
                df["Buyer(%)"] + 
                df["Import tax(%)"] + 
                df["VAT"] + 
                df["Transportation"] + 
                df["Management fee(%)"])

    # L·ª£i nhu·∫≠n = Doanh thu - Chi ph√≠ + Payback
    df["Profit(VND)"] = df["Total price(VND)"] - cost_ops + df["Payback(%)"]
    
    # T√≠nh % L·ª£i nhu·∫≠n
    df["Profit_Pct_Raw"] = df.apply(lambda row: (row["Profit(VND)"] / row["Total price(VND)"] * 100) if row["Total price(VND)"] > 0 else 0, axis=1)
    df["Profit(%)"] = df["Profit_Pct_Raw"].apply(lambda x: f"{x:.1f}%")
    
    # C·∫£nh b√°o
    def set_warning(row):
        if "KH√îNG KH·ªöP" in str(row["C·∫£nh b√°o"]): return row["C·∫£nh b√°o"]
        return "‚ö†Ô∏è LOW" if row["Profit_Pct_Raw"] < 10 else "‚úÖ OK"
    df["C·∫£nh b√°o"] = df.apply(set_warning, axis=1)

    return df

# --- IMPROVED FORMULA PARSER ---
def parse_formula(formula, buying_price, ap_price):
    if not formula: return 0.0
    
    # 1. Normalize: Uppercase and Strip
    s = str(formula).strip().upper()
    
    # 2. Handle '='
    if s.startswith("="): s = s[1:]
    
    # 3. Replace Keywords (Longer first to avoid substrings issue)
    # Handle 'AP PRICE' explicitly before 'AP'
    s = s.replace("AP PRICE", str(ap_price))
    s = s.replace("BUYING PRICE", str(buying_price))
    
    # Handle shorthands
    s = s.replace("AP", str(ap_price))
    s = s.replace("BUY", str(buying_price))
    
    # 4. Cleanup Syntax
    s = s.replace(",", ".").replace("%", "/100").replace("X", "*")
    
    # 5. Filter Unsafe Characters (Only digits, dots, math ops)
    s = re.sub(r'[^0-9.+\-*/()]', '', s)
    
    try: 
        if not s: return 0.0
        return float(eval(s))
    except: return 0.0

# =============================================================================
# 4. GIAO DI·ªÜN CH√çNH
# =============================================================================
t1, t2, t3, t4, t5, t6 = st.tabs(["üìä DASHBOARD", "üì¶ KHO H√ÄNG", "üí∞ B√ÅO GI√Å", "üìë QU·∫¢N L√ù PO", "üöö TRACKING", "‚öôÔ∏è MASTER DATA"])

# --- TAB 1: DASHBOARD ---
with t1:
    if st.button("üîÑ REFRESH DATA"): st.cache_data.clear(); st.rerun()
    db_cust = load_data("db_customer_orders")
    db_supp = load_data("db_supplier_orders")
    rev = db_cust['total_price'].apply(to_float).sum() if not db_cust.empty else 0
    cost = db_supp['total_vnd'].apply(to_float).sum() if not db_supp.empty else 0
    profit = rev - cost 
    c1, c2, c3 = st.columns(3)
    c1.markdown(f"<div class='card-3d bg-sales'><h3>DOANH THU</h3><h1>{fmt_num(rev)}</h1></div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='card-3d bg-cost'><h3>CHI PH√ç NCC</h3><h1>{fmt_num(cost)}</h1></div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='card-3d bg-profit'><h3>L·ª¢I NHU·∫¨N G·ªòP</h3><h1>{fmt_num(profit)}</h1></div>", unsafe_allow_html=True)

# --- TAB 2: KHO H√ÄNG ---
with t2:
    st.subheader("QU·∫¢N L√ù KHO H√ÄNG (Excel Online)")
    c_imp, c_view = st.columns([1, 4])
    
    with c_imp:
        st.markdown("**üì• Import Kho H√†ng**")
        st.caption("Excel c·ªôt A->O")
        st.info("No, Code, Name, Specs, Qty, BuyRMB, TotalRMB, Rate, BuyVND, TotalVND, Leadtime, Supplier, Images, Type, N/U/O/C")
        
        with st.expander("üõ†Ô∏è Reset DB"):
            adm_pass = st.text_input("Pass", type="password", key="adm_inv")
            if st.button("‚ö†Ô∏è X√ìA S·∫†CH"):
                if adm_pass == "admin":
                    supabase.table("crm_purchases").delete().neq("id", 0).execute()
                    st.success("Deleted!"); time.sleep(1); st.rerun()
                else: st.error("Sai Pass!")
        
        up_file = st.file_uploader("Upload Excel", type=["xlsx"], key="inv_up")
            
        if up_file and st.button("üöÄ Import"):
            try:
                wb = load_workbook(up_file, data_only=False); ws = wb.active
                img_map = {}
                for image in getattr(ws, '_images', []):
                    row = image.anchor._from.row + 1
                    buf = io.BytesIO(image._data())
                    cell_specs = ws.cell(row=row, column=4).value 
                    specs_val = safe_str(cell_specs)
                    safe_name = re.sub(r'[\\/*?:"<>|]', "", specs_val).strip()
                    if not safe_name: safe_name = f"NO_SPECS_R{row}"
                    fname = f"{safe_name}.png"
                    link, _ = upload_to_drive_simple(buf, "CRM_PRODUCT_IMAGES", fname)
                    img_map[row] = link
                
                df = pd.read_excel(up_file, header=None, skiprows=1, dtype=str).fillna("")
                records = []
                prog = st.progress(0)
                cols_map = ["no", "item_code", "item_name", "specs", "qty", "buying_price_rmb", 
                            "total_buying_price_rmb", "exchange_rate", "buying_price_vnd", 
                            "total_buying_price_vnd", "leadtime", "supplier_name", "image_path", "type", "nuoc"]

                for i, r in df.iterrows():
                    d = {}
                    for idx, field in enumerate(cols_map):
                        # --- FIX L·ªñI PGRST204: B·ªè qua c·ªôt 'no' v√¨ DB kh√¥ng c√≥ c·ªôt n√†y ---
                        # (N·∫øu b·∫°n ƒë√£ th√™m c·ªôt 'no' b·∫±ng SQL th√¨ c√≥ th·ªÉ b·ªè d√≤ng n√†y, nh∆∞ng ƒë·ªÉ an to√†n c·ª© gi·ªØ logic map)
                        if idx < len(r): d[field] = safe_str(r.iloc[idx])
                        else: d[field] = ""
                    has_data = d['item_code'] or d['item_name'] or d['specs']
                    if has_data:
                        if not d.get('image_path') and (i+2) in img_map: d['image_path'] = img_map[i+2]
                        d['row_order'] = i + 1 
                        d['qty'] = to_float(d.get('qty', 0))
                        d['buying_price_rmb'] = to_float(d['buying_price_rmb'])
                        d['total_buying_price_rmb'] = to_float(d['total_buying_price_rmb'])
                        d['exchange_rate'] = to_float(d['exchange_rate'])
                        d['buying_price_vnd'] = to_float(d['buying_price_vnd'])
                        d['total_buying_price_vnd'] = to_float(d['total_buying_price_vnd'])
                        records.append(d)
                    prog.progress((i + 1) / len(df))
                
                if records:
                    chunk_ins = 100
                    codes = [b['item_code'] for b in records if b['item_code']]
                    if codes: supabase.table("crm_purchases").delete().in_("item_code", codes).execute()
                    
                    # --- FIX: TH·ª¨ INSERT B√åNH TH∆Ø·ªúNG TR∆Ø·ªöC ---
                    try:
                        for k in range(0, len(records), chunk_ins):
                            batch = records[k:k+chunk_ins]
                            # --- FIX L·ªñI 23505: ƒê·ªîI SANG UPSERT ---
                            supabase.table("crm_purchases").upsert(batch).execute()
                    except Exception as e_ins:
                         # N·∫æU L·ªñI DO C·ªòT row_order KH√îNG T·ªíN T·∫†I, X√ìA N√ì ƒêI V√Ä INSERT L·∫†I
                        if "row_order" in str(e_ins):
                            st.warning("‚ö†Ô∏è Database c≈© ch∆∞a c√≥ c·ªôt row_order. ƒêang t·ª± ƒë·ªông b·ªè qua c·ªôt n√†y ƒë·ªÉ insert...")
                            for rec in records:
                                if 'row_order' in rec: del rec['row_order']
                            
                            for k in range(0, len(records), chunk_ins):
                                batch = records[k:k+chunk_ins]
                                # --- FIX L·ªñI 23505: ƒê·ªîI SANG UPSERT ---
                                supabase.table("crm_purchases").upsert(batch).execute()
                        else:
                            raise e_ins

                    st.success(f"‚úÖ ƒê√£ import {len(records)} d√≤ng (ƒë√∫ng th·ª© t·ª± Excel)!")
                    st.cache_data.clear(); time.sleep(1); st.rerun()
            except Exception as e: st.error(f"L·ªói Import: {e}")

    with c_view:
        # Load data ƒë√£ fix sort
        df_pur = load_data("crm_purchases", order_by="row_order", ascending=True) 
        cols_to_drop = ['created_at', 'row_order']
        df_pur = df_pur.drop(columns=[c for c in cols_to_drop if c in df_pur.columns], errors='ignore')

        # --- FIX: ƒê·ªîI TH·ª® T·ª∞ C·ªòT ƒê·ªÇ C·ªòT 'no' HO·∫∂C 'No' L√äN ƒê·∫¶U ---
        current_cols = df_pur.columns.tolist()
        # T√¨m c·ªôt 'no' ho·∫∑c 'No'
        no_col = next((c for c in current_cols if c.lower() == 'no'), None)
        
        if no_col:
            # X√≥a kh·ªèi v·ªã tr√≠ c≈©
            current_cols.remove(no_col)
            # Ch√®n v√†o ƒë·∫ßu
            current_cols.insert(0, no_col)
            # √Åp d·ª•ng
            df_pur = df_pur[current_cols]

        search = st.text_input("üîç T√¨m ki·∫øm (Name, Code, Specs...)", key="search_pur")
        if not df_pur.empty:
            if search:
                mask = df_pur.astype(str).apply(lambda x: x.str.contains(search, case=False, na=False)).any(axis=1)
                df_pur = df_pur[mask]
            
            cols_money = ["buying_price_vnd", "total_buying_price_vnd", "buying_price_rmb", "total_buying_price_rmb"]
            for c in cols_money:
                if c in df_pur.columns: df_pur[c] = df_pur[c].apply(fmt_num)

            st.dataframe(
                df_pur, 
                column_config={
                    "image_path": st.column_config.ImageColumn("Images"),
                    "item_code": st.column_config.TextColumn("Code", width="medium"),
                    "item_name": st.column_config.TextColumn("Name", width="medium"),
                    "specs": st.column_config.TextColumn("Specs", width="large"),
                    "buying_price_vnd": st.column_config.TextColumn("Buying (VND)"),
                    "total_buying_price_vnd": st.column_config.TextColumn("Total (VND)"),
                    "buying_price_rmb": st.column_config.TextColumn("Buying (RMB)"),
                    "total_buying_price_rmb": st.column_config.TextColumn("Total (RMB)"),
                    "qty": st.column_config.NumberColumn("Qty", format="%d"),
                }, 
                use_container_width=True, height=700, hide_index=True
            )
        else: st.info("Kho h√†ng tr·ªëng.")

# --- TAB 3: B√ÅO GI√Å ---
with t3:
    if 'quote_df' not in st.session_state: st.session_state.quote_df = pd.DataFrame()
    
    # ------------------ TRA C·ª®U L·ªäCH S·ª¨ ------------------
    with st.expander("üîé TRA C·ª®U & TR·∫†NG TH√ÅI B√ÅO GI√Å", expanded=False):
        c_src1, c_src2 = st.columns(2)
        search_kw = c_src1.text_input("Nh·∫≠p t·ª´ kh√≥a (T√™n Kh√°ch, Quote No, Code, Name, Date)", help="T√¨m ki·∫øm trong l·ªãch s·ª≠")
        up_src = c_src2.file_uploader("Ho·∫∑c Import Excel ki·ªÉm tra", type=["xlsx"], key="src_up")
        
        if st.button("Ki·ªÉm tra tr·∫°ng th√°i"):
            df_hist = load_data("crm_shared_history")
            df_po = load_data("db_customer_orders")
            df_items = load_data("crm_purchases") 

            item_map = {}
            if not df_items.empty:
                for r in df_items.to_dict('records'):
                    k = clean_key(r['item_code'])
                    item_map[k] = f"{safe_str(r['item_name'])} {safe_str(r['specs'])}"

            po_map = {}
            if not df_po.empty:
                for r in df_po.to_dict('records'):
                    k = f"{clean_key(r['customer'])}_{clean_key(r['item_code'])}"
                    po_map[k] = r['po_number']

            results = []
            if search_kw and not df_hist.empty:
                def check_row(row):
                    kw = search_kw.lower()
                    if kw in str(row.get('customer','')).lower(): return True
                    if kw in str(row.get('quote_no','')).lower(): return True
                    if kw in str(row.get('item_code','')).lower(): return True
                    if kw in str(row.get('date','')).lower(): return True
                    code = clean_key(row['item_code'])
                    info = item_map.get(code, "").lower()
                    if kw in info: return True
                    return False
                
                mask = df_hist.apply(check_row, axis=1)
                found = df_hist[mask]
                for _, r in found.iterrows():
                    key = f"{clean_key(r['customer'])}_{clean_key(r['item_code'])}"
                    po_found = po_map.get(key, "")
                    code_info = item_map.get(clean_key(r['item_code']), "")
                    results.append({
                        "Tr·∫°ng th√°i": "‚úÖ ƒê√£ b√°o gi√°", "Customer": r['customer'], "Date": r['date'],
                        "Item Code": r['item_code'], "Info": code_info, 
                        "Unit Price": fmt_float_2(r['unit_price']),
                        "Quote No": r['quote_no'], "PO No": po_found if po_found else "---"
                    })
            
            if up_src:
                try:
                    df_check = pd.read_excel(up_src, dtype=str).fillna("")
                    cols_check = {clean_key(c): c for c in df_check.columns}
                    for i, r in df_check.iterrows():
                        code = ""; name = ""; specs = ""
                        for k, col in cols_check.items():
                            if "code" in k: code = safe_str(r[col])
                            elif "name" in k: name = safe_str(r[col])
                            elif "specs" in k: specs = safe_str(r[col])
                        match = pd.DataFrame()
                        if not df_hist.empty:
                            if code: match = df_hist[df_hist['item_code'].str.contains(code, case=False, na=False)]
                        if not match.empty:
                            for _, m in match.iterrows():
                                key = f"{clean_key(m['customer'])}_{clean_key(m['item_code'])}"
                                po_found = po_map.get(key, "")
                                results.append({
                                    "Tr·∫°ng th√°i": "‚úÖ ƒê√£ b√°o gi√°", "Customer": m['customer'], "Date": m['date'],
                                    "Item Code": m['item_code'], "Info": item_map.get(clean_key(m['item_code']), ""),
                                    "Unit Price": fmt_float_2(m['unit_price']), "Quote No": m['quote_no'], "PO No": po_found
                                })
                        else:
                            results.append({
                                "Tr·∫°ng th√°i": "‚ùå Ch∆∞a b√°o gi√°", "Item Code": code, "Customer": "---", 
                                "Date": "---", "Unit Price": "---", "Quote No": "---", "PO No": "---"
                            })
                except Exception as e: st.error(f"L·ªói file: {e}")

            if results: st.dataframe(pd.DataFrame(results), use_container_width=True)
            else: st.info("Kh√¥ng t√¨m th·∫•y k·∫øt qu·∫£.")

    with st.expander("üìÇ XEM CHI TI·∫æT FILE L·ªäCH S·ª¨ (COST & L·ª¢I NHU·∫¨N)", expanded=False):
        df_hist_idx = load_data("crm_shared_history", order_by="date")
        if not df_hist_idx.empty:
            df_hist_idx['display'] = df_hist_idx.apply(lambda x: f"{x['date']} | {x['customer']} | Quote: {x['quote_no']}", axis=1)
            unique_quotes = df_hist_idx['display'].unique()
            filtered_quotes = unique_quotes
            if search_kw: filtered_quotes = [q for q in unique_quotes if search_kw.lower() in q.lower()]
            sel_quote_hist = st.selectbox("Ch·ªçn b√°o gi√° c≈© ƒë·ªÉ xem chi ti·∫øt:", [""] + list(filtered_quotes))
            
            if sel_quote_hist:
                parts = sel_quote_hist.split(" | ")
                if len(parts) >= 3:
                    q_no = parts[2].replace("Quote: ", "").strip()
                    cust = parts[1].strip()
                    
                    # --- HOTFIX: FORCE RELOAD CONFIG FROM EXCEL FALLBACK OR DB ---
                    # Check if new quote selected to force RERUN
                    if 'loaded_quote_id' not in st.session_state: st.session_state.loaded_quote_id = None
                    
                    hist_config_row = df_hist_idx[
                        (df_hist_idx['quote_no'] == q_no) & 
                        (df_hist_idx['customer'] == cust)
                    ].iloc[0] if not df_hist_idx.empty else None
                    
                    config_loaded = {}
                    
                    # 1. Try DB
                    if hist_config_row is not None and 'config_data' in hist_config_row and hist_config_row['config_data']:
                        try:
                            config_loaded = json.loads(hist_config_row['config_data'])
                        except: pass
                    
                    # 2. If DB empty, Try Drive (Fallback)
                    if not config_loaded:
                          cfg_search_name = f"CONFIG_{q_no}_{cust}"
                          fid_cfg, _, _ = search_file_in_drive_by_name(cfg_search_name)
                          if fid_cfg:
                              fh_cfg = download_from_drive(fid_cfg)
                              if fh_cfg:
                                  try:
                                      df_cfg = pd.read_excel(fh_cfg)
                                      if not df_cfg.empty:
                                          config_loaded = df_cfg.iloc[0].to_dict()
                                  except: pass

                    # 3. Apply Config
                    if config_loaded:
                        st.info(f"üìä **C·∫§U H√åNH CHI PH√ç (ƒê√É LOAD):** "
                                f"End User: {config_loaded.get('end')}% | Buyer: {config_loaded.get('buy')}% | "
                                f"Tax: {config_loaded.get('tax')}% | VAT: {config_loaded.get('vat')}% | "
                                f"Payback: {config_loaded.get('pay')}% | Mgmt: {config_loaded.get('mgmt')}% | "
                                f"Trans: {fmt_num(config_loaded.get('trans'))}")
                        
                        # Trigger RERUN if switching quotes to update widgets
                        if sel_quote_hist != st.session_state.loaded_quote_id:
                            keys_load = ["end", "buy", "tax", "vat", "pay", "mgmt", "trans"]
                            for k in keys_load:
                                val_str = str(config_loaded.get(k, 0))
                                st.session_state[f"pct_{k}"] = val_str
                                st.session_state[f"input_{k}"] = val_str # Force Widget Key
                            
                            st.session_state.loaded_quote_id = sel_quote_hist
                            st.toast("‚úÖ ƒê√£ load c·∫•u h√¨nh th√†nh c√¥ng!", icon="‚úÖ")
                            time.sleep(0.5)
                            st.rerun()
                    else:
                        st.warning("‚ö†Ô∏è B√°o gi√° n√†y ƒë∆∞·ª£c t·∫°o t·ª´ phi√™n b·∫£n c≈©, ch∆∞a l∆∞u c·∫•u h√¨nh chi ph√≠.")

                    search_name = f"HIST_{q_no}_{cust}"
                    fid, fname, pid = search_file_in_drive_by_name(search_name)
                    if pid:
                          folder_link = f"https://drive.google.com/drive/folders/{pid}"
                          st.markdown(f"üëâ **[M·ªü Folder ch·ª©a file n√†y tr√™n Google Drive]({folder_link})**", unsafe_allow_html=True)
                    if fid and st.button(f"T·∫£i file chi ti·∫øt: {fname}"):
                          fh = download_from_drive(fid)
                          if fh:
                              try:
                                  df_csv = pd.read_csv(fh, encoding='utf-8-sig', on_bad_lines='skip')
                                  st.success("ƒê√£ t·∫£i xong!")
                                  st.dataframe(df_csv, use_container_width=True)
                              except Exception as e: st.error(f"L·ªói ƒë·ªçc file CSV: {e}")
                          else: st.error("Kh√¥ng t·∫£i ƒë∆∞·ª£c file.")
                    elif not fid: st.warning(f"Kh√¥ng t√¨m th·∫•y file chi ti·∫øt tr√™n Drive (HIST_{q_no}...).")
        else: st.info("Ch∆∞a c√≥ l·ªãch s·ª≠.")

    st.divider()
    st.subheader("T√çNH TO√ÅN & L√ÄM B√ÅO GI√Å")
    
    c1, c2, c3 = st.columns([2, 2, 1])
    cust_db = load_data("crm_customers")
    cust_list = cust_db["short_name"].tolist() if not cust_db.empty else []
    cust_name = c1.selectbox("Ch·ªçn Kh√°ch H√†ng", [""] + cust_list)
    quote_no = c2.text_input("S·ªë B√°o Gi√°", key="q_no")
    
    c3.markdown('<div class="dark-btn">', unsafe_allow_html=True)
    if c3.button("üîÑ Reset Quote"): 
        st.session_state.quote_df = pd.DataFrame()
        st.session_state.show_review = False 
        for k in ["end","buy","tax","vat","pay","mgmt","trans"]:
             if f"pct_{k}" in st.session_state: del st.session_state[f"pct_{k}"]
        st.rerun()
    c3.markdown('</div>', unsafe_allow_html=True)

    with st.expander("C·∫•u h√¨nh chi ph√≠ (%) & V·∫≠n chuy·ªÉn", expanded=True):
        cols = st.columns(7)
        keys = ["end", "buy", "tax", "vat", "pay", "mgmt", "trans"]
        params = {}
        for i, k in enumerate(keys):
            default_val = st.session_state.get(f"pct_{k}", "0")
            # --- WIDGET INPUT ---
            # Quan tr·ªçng: key=f"input_{k}" ƒë·ªÉ kh·ªõp v·ªõi logic load l·ªãch s·ª≠
            val = cols[i].text_input(k.upper(), value=default_val, key=f"input_{k}")
            st.session_state[f"pct_{k}"] = val
            params[k] = to_float(val)

    # MATCHING
    cf1, cf2 = st.columns([1, 2])
    rfq = cf1.file_uploader("Upload RFQ (xlsx)", type=["xlsx"])
    if rfq and cf2.button("üîç Matching"):
        st.session_state.quote_df = pd.DataFrame()
        db = load_data("crm_purchases")
        if db.empty: st.error("Kho r·ªóng!")
        else:
            db_records = db.to_dict('records')
            df_rfq = pd.read_excel(rfq, dtype=str).fillna("")
            res = []
            cols_found = {clean_key(c): c for c in df_rfq.columns}
            
            # --- FIX: PROGRESS BAR CHO MATCHING ---
            prog_match = st.progress(0)
            status_match = st.empty()
            total_rows = len(df_rfq)
            
            for i, r in df_rfq.iterrows():
                # --- UPDATE PROGRESS ---
                pct = (i + 1) / total_rows
                prog_match.progress(pct)
                status_match.text(f"ƒêang x·ª≠ l√Ω: {int(pct*100)}%")

                def get_val(keywords):
                    for k in keywords:
                        real_col = cols_found.get(k)
                        if real_col: return safe_str(r[real_col])
                    return ""

                # 1. L·∫§Y D·ªÆ LI·ªÜU T·ª™ EXCEL (SOURCE OF TRUTH)
                code_excel = get_val(["item code", "code", "m√£", "part number"])
                name_excel = get_val(["item name", "name", "t√™n", "description"])
                specs_excel = get_val(["specs", "quy c√°ch", "th√¥ng s·ªë"])
                qty_raw = get_val(["q'ty", "qty", "quantity", "s·ªë l∆∞·ª£ng"])
                qty = to_float(qty_raw) if qty_raw else 1.0

                # 2. MATCHING LOGIC (Kh·ªõp 3 th√¥ng s·ªë: Code, Name, Specs)
                match = None
                warning_msg = ""
                
                candidates = [
                    rec for rec in db_records 
                    if clean_key(rec['item_code']) == clean_key(code_excel)
                    and clean_key(rec['item_name']) == clean_key(name_excel)
                    and clean_key(rec['specs']) == clean_key(specs_excel)
                ]

                if candidates:
                    match = candidates[0]
                else:
                    warning_msg = "‚ö†Ô∏è KH√îNG KH·ªöP DATA"

                if match:
                    buy_rmb = to_float(match.get('buying_price_rmb', 0))
                    buy_vnd = to_float(match.get('buying_price_vnd', 0))
                    ex_rate = to_float(match.get('exchange_rate', 0))
                    supplier = match.get('supplier_name', '')
                    image = match.get('image_path', '')
                    leadtime = match.get('leadtime', '')
                else:
                    buy_rmb = 0; buy_vnd = 0; ex_rate = 0
                    supplier = ""; image = ""; leadtime = ""

                # --- L·∫§Y GI√Å TR·ªä GLOBAL PARAM HI·ªÜN T·∫†I (ƒê·ªÇ KH·ªûI T·∫†O) ---
                p_end = params['end']/100; p_buy = params['buy']/100
                p_tax = params['tax']/100; p_vat = params['vat']/100
                p_pay = params['pay']/100; p_mgmt = params['mgmt']/100
                v_trans = params['trans']

                # T√≠nh to√°n s∆° b·ªô cho Import Tax (v√¨ c√≥ gi√° mua r·ªìi)
                val_import_tax = (buy_vnd * qty) * p_tax

                item = {
                    "Select": False, "No": i+1, "C·∫£nh b√°o": warning_msg, 
                    "Item code": code_excel, "Item name": name_excel, "Specs": specs_excel, 
                    "Q'ty": qty, 
                    "Buying price(RMB)": fmt_float_2(buy_rmb), 
                    "Total buying price(rmb)": fmt_float_2(buy_rmb * qty),
                    "Exchange rate": fmt_float_2(ex_rate), 
                    "Buying price(VND)": fmt_float_2(buy_vnd), 
                    "Total buying price(VND)": fmt_float_2(buy_vnd * qty),
                    
                    "AP price(VND)": "0.00", "AP total price(VND)": "0.00", 
                    "Unit price(VND)": "0.00", "Total price(VND)": "0.00",
                    "GAP": "0.00",
                    
                    # --- KH·ªûI T·∫†O GI√Å TR·ªä T·ª™ GLOBAL CONFIG ---
                    "End user(%)": "0.00",       
                    "Buyer(%)": "0.00",          
                    "Import tax(%)": fmt_float_2(val_import_tax), # T√≠nh lu√¥n v√¨ ƒë√£ c√≥ gi√° mua
                    "VAT": "0.00",               
                    "Transportation": fmt_num(v_trans), # Kh·ªüi t·∫°o b·∫±ng Global
                    "Management fee(%)": "0.00",
                    "Payback(%)": "0.00",        
                    # ----------------------------------------

                    "Profit(VND)": "0.00", "Profit(%)": "0.0%",
                    "Supplier": supplier, "Image": image, "Leadtime": leadtime
                }
                res.append(item)
            
            # --- DONE PROGRESS ---
            prog_match.progress(100)
            status_match.text("Done! Ho√†n t·∫•t Matching.")
            time.sleep(1)
            status_match.empty() # ·∫®n text Done ƒë·ªÉ s·∫°ch giao di·ªán

            st.session_state.quote_df = pd.DataFrame(res)
    
    # --- FORMULA BUTTONS (ONE CLICK FIX) ---
    c_form1, c_form2 = st.columns(2)
    with c_form1:
        ap_f = st.text_input("Formula AP (vd: =BUY*1.1)", key="f_ap")
        st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
        if st.button("Apply AP Price"):
            if not st.session_state.quote_df.empty:
                for idx, row in st.session_state.quote_df.iterrows():
                    buy = to_float(row["Buying price(VND)"])
                    ap = to_float(row["AP price(VND)"])
                    new_ap = parse_formula(ap_f, buy, ap)
                    st.session_state.quote_df.at[idx, "AP price(VND)"] = fmt_float_2(new_ap)
                st.session_state.quote_df = recalculate_quote_logic(st.session_state.quote_df, params)
                st.rerun() 
        st.markdown('</div>', unsafe_allow_html=True)
    with c_form2:
        unit_f = st.text_input("Formula Unit (vd: =AP*1.2)", key="f_unit")
        st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
        if st.button("Apply Unit Price"):
            if not st.session_state.quote_df.empty:
                for idx, row in st.session_state.quote_df.iterrows():
                    buy = to_float(row["Buying price(VND)"])
                    ap = to_float(row["AP price(VND)"])
                    new_unit = parse_formula(unit_f, buy, ap)
                    st.session_state.quote_df.at[idx, "Unit price(VND)"] = fmt_float_2(new_unit)
                st.session_state.quote_df = recalculate_quote_logic(st.session_state.quote_df, params)
                st.rerun() 
        st.markdown('</div>', unsafe_allow_html=True)
    
    if not st.session_state.quote_df.empty:
        # --- FIX: ENSURE SELECT COLUMN EXISTS ---
        if "Select" not in st.session_state.quote_df.columns:
            st.session_state.quote_df.insert(0, "Select", False)

        # REAL-TIME CALCULATION BEFORE DISPLAY (Fixes Transportation lag)
        st.session_state.quote_df = recalculate_quote_logic(st.session_state.quote_df, params)

        # --- FIX: ƒê·ªîI TH·ª® T·ª∞ C·ªòT ƒê·ªÇ C·ªòT 'Select' -> 'No' L√äN ƒê·∫¶U ---
        cols_order = ["Select", "No", "C·∫£nh b√°o"] + [c for c in st.session_state.quote_df.columns if c not in ["Select", "No", "C·∫£nh b√°o"]]
        st.session_state.quote_df = st.session_state.quote_df[cols_order]

        cols_to_hide = ["Image", "Profit_Pct_Raw"]
        df_show = st.session_state.quote_df.drop(columns=[c for c in cols_to_hide if c in st.session_state.quote_df.columns], errors='ignore')

        # --- ADD TOTAL ROW LOGIC ---
        df_display = df_show.copy()
        
        # Calculate sums for relevant columns (UPDATED LIST)
        cols_to_sum = ["Q'ty", "Buying price(RMB)", "Total buying price(rmb)", 
                       "Buying price(VND)", "Total buying price(VND)", 
                       "AP price(VND)", "AP total price(VND)", 
                       "Unit price(VND)", "Total price(VND)", 
                       "GAP", "End user(%)", "Buyer(%)", "Import tax(%)", 
                       "VAT", "Transportation", "Management fee(%)", "Payback(%)", "Profit(VND)"]
        
        total_row = {"Select": False, "No": "TOTAL", "C·∫£nh b√°o": "", "Item code": "", "Item name": "", "Specs": "", "Q'ty": 0}
        
        # Calculate sums for ALL requested columns (ON FLOAT DATA)
        for c in cols_to_sum:
            if c in df_display.columns:
                # --- FIX: DATA TYPE CONVERSION FOR SUM ---
                # Chuy·ªÉn ƒë·ªïi d·ªØ li·ªáu v·ªÅ float tr∆∞·ªõc khi sum (tr√°nh l·ªói c·ªông chu·ªói n·∫øu ƒëang l√† string)
                total_val = df_display[c].apply(to_float).sum()
                total_row[c] = fmt_float_2(total_val)
        
        # Recalculate Profit % for Total Row
        t_profit = to_float(total_row.get("Profit(VND)", 0))
        t_price = to_float(total_row.get("Total price(VND)", 0))
        t_pct = (t_profit / t_price * 100) if t_price > 0 else 0
        total_row["Profit(%)"] = f"{t_pct:.1f}%"
        
        # Append Total Row to dataframe for display
        df_display = pd.concat([df_display, pd.DataFrame([total_row])], ignore_index=True)

        # --- N√öT CH·ª®C NƒÇNG M·ªöI: √ÅP D·ª§NG L·∫†I CONFIG ---
        st.markdown("---")
        c_tool1, c_tool2 = st.columns([1, 3])
        with c_tool1:
            if st.button("‚ö° √ÅP D·ª§NG GLOBAL CONFIG", help="T√≠nh l·∫°i to√†n b·ªô c√°c c·ªôt Tax, VAT, Fee... theo % Global ƒëang nh·∫≠p ·ªü tr√™n."):
                if not st.session_state.quote_df.empty:
                    # L·∫•y params hi·ªán t·∫°i
                    p_end = params['end']/100; p_buy = params['buy']/100
                    p_tax = params['tax']/100; p_vat = params['vat']/100
                    p_pay = params['pay']/100; p_mgmt = params['mgmt']/100
                    v_trans = params['trans']
                    
                    # Duy·ªát qua t·ª´ng d√≤ng v√† t√≠nh l·∫°i gi√° tr·ªã (L∆ØU D·∫†NG S·ªê FLOAT)
                    for idx, row in st.session_state.quote_df.iterrows():
                        # L·∫•y c√°c gi√° tr·ªã c∆° s·ªü (S·ªë th·ª±c)
                        val_ap_total = to_float(row["AP total price(VND)"])
                        val_total = to_float(row["Total price(VND)"])
                        val_buy_total = to_float(row["Total buying price(VND)"])
                        val_gap = to_float(row["GAP"])
                        
                        # √Åp d·ª•ng c√¥ng th·ª©c Global - L∆ØU FLOAT V√ÄO SESSION STATE
                        st.session_state.quote_df.at[idx, "End user(%)"] = val_ap_total * p_end
                        st.session_state.quote_df.at[idx, "Buyer(%)"] = val_total * p_buy
                        st.session_state.quote_df.at[idx, "Import tax(%)"] = val_buy_total * p_tax
                        st.session_state.quote_df.at[idx, "VAT"] = val_total * p_vat
                        st.session_state.quote_df.at[idx, "Management fee(%)"] = val_total * p_mgmt
                        st.session_state.quote_df.at[idx, "Payback(%)"] = val_gap * p_pay
                        st.session_state.quote_df.at[idx, "Transportation"] = v_trans # L∆∞u s·ªë th·ª±c
                    
                    st.toast("‚úÖ ƒê√£ √°p d·ª•ng Global Config cho to√†n b·ªô b·∫£ng!", icon="‚ö°")
                    st.rerun()

        # --- DATA EDITOR (FIX INDENTATION) ---
        # Convert sang string format ƒë·∫πp ƒë·ªÉ hi·ªÉn th·ªã (CH·ªà DISPLAY)
        cols_display_fmt = ["Exchange rate", "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", "Transportation", "Management fee(%)", "Payback(%)"]
        for c in cols_display_fmt:
            if c in df_display.columns:
                # --- FIX QUAN TR·ªåNG: CH·ªà FORMAT N·∫æU KH√îNG PH·∫¢I D√íNG TOTAL ---
                # D√≤ng Total ƒë√£ ƒë∆∞·ª£c format ·ªü tr√™n, c√°c d√≤ng d·ªØ li·ªáu c·∫ßn format ƒë·ªÉ hi·ªÉn th·ªã
                df_display[c] = df_display.apply(lambda r: fmt_num(r[c]) if r['No'] != "TOTAL" else r[c], axis=1)

        edited_df = st.data_editor(
            df_display,
            column_config={
                "Select": st.column_config.CheckboxColumn("‚úÖ", width="small"),
                "Buying price(RMB)": st.column_config.TextColumn("Buying(RMB)"),
                "Exchange rate": st.column_config.TextColumn("Rate", width="small"),
                "Buying price(VND)": st.column_config.TextColumn("Buying(VND)"),
                
                # --- C√ÅC C·ªòT M·ªöI CHO PH√âP S·ª¨A ---
                "End user(%)": st.column_config.TextColumn("EndUser(VNƒê)"),
                "Buyer(%)": st.column_config.TextColumn("Buyer(VNƒê)"),
                "Import tax(%)": st.column_config.TextColumn("Tax(VNƒê)"),
                "VAT": st.column_config.TextColumn("VAT(VNƒê)"),
                "Transportation": st.column_config.TextColumn("Trans(VNƒê)"),
                "Management fee(%)": st.column_config.TextColumn("Mgmt(VNƒê)"),
                "Payback(%)": st.column_config.TextColumn("Payback(VNƒê)"),
                # --------------------------------
                
                "C·∫£nh b√°o": st.column_config.TextColumn("C·∫£nh b√°o", width="small", disabled=True),
                "Q'ty": st.column_config.NumberColumn("Q'ty", format="%d"),
            },
            use_container_width=True, height=600, key="main_editor",
            hide_index=True 
        )
        
        # --- SYNC EDITS BACK (Exclude Total Row) ---
        df_data_only = edited_df[edited_df["No"] != "TOTAL"]
        
        # Check if data changed to trigger recalculation next run
        # Update main dataframe with edited values (mapped back)
        data_changed = False
        for idx, row in df_data_only.iterrows():
             if idx < len(st.session_state.quote_df):
                 for c in df_data_only.columns:
                     if c in st.session_state.quote_df.columns:
                         old_val = st.session_state.quote_df.at[idx, c]
                         new_val = row[c]
                         
                         # --- FIX FLICKERING: COMPARE VALUES NOT STRINGS ---
                         # C√°c c·ªôt s·ªë li·ªáu (Money/Percent)
                         numeric_cols = ["Q'ty", "Buying price(RMB)", "Exchange rate", "Buying price(VND)", 
                                         "End user(%)", "Buyer(%)", "Import tax(%)", "VAT", 
                                         "Transportation", "Management fee(%)", "Payback(%)"]
                         
                         if c in numeric_cols:
                             # So s√°nh s·ªë (tr√°nh l·ªói 100000.0 != 100,000)
                             if abs(to_float(old_val) - to_float(new_val)) > 0.001:
                                 st.session_state.quote_df.at[idx, c] = new_val # L∆∞u gi√° tr·ªã m·ªõi v√†o
                                 data_changed = True
                         else:
                             # C√°c c·ªôt Text/Checkbox
                             if str(old_val) != str(new_val):
                                 st.session_state.quote_df.at[idx, c] = new_val
                                 data_changed = True
        
        if data_changed:
            st.rerun()

        # --- QUICK TOOLBAR LOGIC ---
        # Get selected rows from the edited dataframe (excluding Total row if selected)
        selected_rows = df_data_only[df_data_only["Select"] == True]
        
        if not selected_rows.empty:
            st.info(f"ƒêang ch·ªçn {len(selected_rows)} d√≤ng.")
            tb_col1, tb_col2, tb_col3, tb_col4 = st.columns(4)
            
            with tb_col1:
                if st.button("üóëÔ∏è DELETE"):
                    # Get indices of selected rows
                    indices_to_drop = selected_rows.index
                    st.session_state.quote_df = st.session_state.quote_df.drop(indices_to_drop).reset_index(drop=True)
                    # Re-assign No.
                    st.session_state.quote_df["No"] = st.session_state.quote_df.index + 1
                    st.rerun()
            
            with tb_col2:
                # Download CSV of selected
                csv = selected_rows.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="‚¨áÔ∏è DOWNLOAD",
                    data=csv,
                    file_name="selected_items.csv",
                    mime="text/csv",
                )

            with tb_col3:
                if st.button("‚õ∂ FULL SCREEN"):
                    st.toast("Ch·∫ø ƒë·ªô to√†n m√†n h√¨nh ƒë√£ ƒë∆∞·ª£c k√≠ch ho·∫°t (Gi·∫£ l·∫≠p)", icon="üñ•Ô∏è")
            
            with tb_col4:
                if st.button("üëÅÔ∏è HIDE"):
                      # Remove from view but maybe keep in backend? For now, simpler to just uncheck.
                      # Or assume hide means remove selection
                      for idx in selected_rows.index:
                          st.session_state.quote_df.at[idx, "Select"] = False
                      st.rerun()

        # --- VIEW TOTAL PRICE (FEATURE ADDED) ---
        total_q = st.session_state.quote_df["Total price(VND)"].apply(to_float).sum()
        st.markdown(f'<div class="total-view">üí∞ T·ªîNG GI√Å TR·ªä B√ÅO GI√Å (TOTAL VIEW): {fmt_float_2(total_q)} VND</div>', unsafe_allow_html=True)

        st.divider()
        c_rev, c_sv = st.columns([1, 1])
        with c_rev:
            st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
            if st.button("üîç REVIEW B√ÅO GI√Å"): st.session_state.show_review = True
            st.markdown('</div>', unsafe_allow_html=True)
        
        if st.session_state.get('show_review', False):
            st.write("### üìã B·∫¢NG REVIEW")
            cols_review = ["No", "Item code", "Item name", "Specs", "Q'ty", "Unit price(VND)", "Total price(VND)", "Leadtime"]
            valid_cols = [c for c in cols_review if c in st.session_state.quote_df.columns]
            
            # --- REVIEW TABLE WITH TOTAL ROW ---
            df_review = st.session_state.quote_df[valid_cols].copy()
            
            # 1. T√≠nh to√°n Total tr√™n d·ªØ li·ªáu g·ªëc (s·ªë th·ª±c)
            total_qty = df_review["Q'ty"].apply(to_float).sum() if "Q'ty" in df_review.columns else 0
            total_unit = df_review["Unit price(VND)"].apply(to_float).sum() if "Unit price(VND)" in df_review.columns else 0
            total_price = df_review["Total price(VND)"].apply(to_float).sum() if "Total price(VND)" in df_review.columns else 0

            # 2. Format c√°c d√≤ng d·ªØ li·ªáu TR∆Ø·ªöC khi th√™m Total
            if "Unit price(VND)" in df_review.columns:
                 df_review["Unit price(VND)"] = df_review["Unit price(VND)"].apply(fmt_num)
            if "Total price(VND)" in df_review.columns:
                 df_review["Total price(VND)"] = df_review["Total price(VND)"].apply(fmt_num)
            
            # 3. T·∫°o d√≤ng Total (ƒê√£ format s·∫µn)
            total_review = {
                "No": "TOTAL", 
                "Item code": "", "Item name": "", "Specs": "", "Leadtime": "",
                "Q'ty": total_qty,
                "Unit price(VND)": fmt_float_2(total_unit), # Format t·∫°i ƒë√¢y lu√¥n
                "Total price(VND)": fmt_float_2(total_price) # Format t·∫°i ƒë√¢y lu√¥n
            }
            
            # 4. G·ªôp v√†o b·∫£ng
            df_review = pd.concat([df_review, pd.DataFrame([total_review])], ignore_index=True)
            
            st.dataframe(df_review, use_container_width=True, hide_index=True)
            
            # Show Total in Review as well
            st.markdown(f'<div class="total-view">üí∞ T·ªîNG C·ªòNG: {fmt_float_2(total_q)} VND</div>', unsafe_allow_html=True)
            
            st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
            if st.button("üì§ XU·∫§T B√ÅO GI√Å (Excel)"):
                if not cust_name: st.error("Ch∆∞a ch·ªçn kh√°ch h√†ng!")
                else:
                    try:
                        df_tmpl = load_data("crm_templates")
                        match_tmpl = df_tmpl[df_tmpl['template_name'].astype(str).str.contains("AAA-QUOTATION", case=False, na=False)]
                        if match_tmpl.empty: st.error("Kh√¥ng t√¨m th·∫•y template 'AAA-QUOTATION'!")
                        else:
                            tmpl_id = match_tmpl.iloc[0]['file_id']
                            fh = download_from_drive(tmpl_id)
                            if not fh: st.error("L·ªói t·∫£i template!")
                            else:
                                wb = load_workbook(fh); ws = wb.active
                                # --- FIX: IMPORT T·ª™ D√íNG 11 (Start row = 11) ---
                                start_row = 11
                                first_leadtime = st.session_state.quote_df.iloc[0]['Leadtime'] if not st.session_state.quote_df.empty else ""
                                ws['H8'] = safe_str(first_leadtime)
                                for idx, row in st.session_state.quote_df.iterrows():
                                    r = start_row + idx
                                    ws[f'A{r}'] = row['No']
                                    ws[f'C{r}'] = row['Item code']
                                    ws[f'D{r}'] = row['Item name']
                                    ws[f'E{r}'] = row['Specs']
                                    ws[f'F{r}'] = to_float(row["Q'ty"])
                                    ws[f'G{r}'] = to_float(row["Unit price(VND)"])
                                    ws[f'H{r}'] = to_float(row["Total price(VND)"])
                                out = io.BytesIO(); wb.save(out); out.seek(0)
                                curr_year = datetime.now().strftime("%Y")
                                curr_month = datetime.now().strftime("%b").upper()
                                fname = f"QUOTE_{quote_no}_{cust_name}_{int(time.time())}.xlsx"
                                path_list = ["QUOTATION_HISTORY", cust_name, curr_year, curr_month]
                                lnk, _ = upload_to_drive_structured(out, path_list, fname)
                                st.success(f"‚úÖ ƒê√£ xu·∫•t b√°o gi√°: {fname}")
                                st.markdown(f"üìÇ [M·ªü Folder]({lnk})", unsafe_allow_html=True)
                                st.download_button(label="üì• T·∫£i File V·ªÅ M√°y", data=out, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    except Exception as e: st.error(f"L·ªói xu·∫•t Excel: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

        with c_sv:
            st.markdown('<div class="dark-btn">', unsafe_allow_html=True)
            if st.button("üíæ L∆ØU L·ªäCH S·ª¨ (QUAN TR·ªåNG ƒê·ªÇ L√ÄM PO)"):
                if cust_name:
                    # 1. CLEAN PARAMS BEFORE JSON DUMP (AVOID NaN IN CONFIG)
                    clean_params = {}
                    for k, v in params.items():
                        if isinstance(v, float) and (np.isnan(v) or np.isinf(v)): clean_params[k] = 0.0
                        else: clean_params[k] = v
                    config_json = json.dumps(clean_params) 
                    
                    recs = []
                    for r in st.session_state.quote_df.to_dict('records'):
                        # --- FIX: DATA CLEANING (NaN -> 0.0) ---
                        val_qty = to_float(r["Q'ty"])
                        val_unit = to_float(r["Unit price(VND)"])
                        val_total = to_float(r["Total price(VND)"])
                        val_profit = to_float(r["Profit(VND)"])
                        
                        # Ensure no NaNs exist (Supabase API Error fix)
                        if np.isnan(val_qty) or np.isinf(val_qty): val_qty = 0.0
                        if np.isnan(val_unit) or np.isinf(val_unit): val_unit = 0.0
                        if np.isnan(val_total) or np.isinf(val_total): val_total = 0.0
                        if np.isnan(val_profit) or np.isinf(val_profit): val_profit = 0.0

                        recs.append({
                            "history_id": f"{cust_name}_{int(time.time())}", "date": datetime.now().strftime("%Y-%m-%d"),
                            "quote_no": quote_no, "customer": cust_name,
                            "item_code": r["Item code"], "qty": val_qty,
                            "unit_price": val_unit,
                            "total_price_vnd": val_total,
                            "profit_vnd": val_profit,
                            "config_data": config_json 
                        })
                    
                    try:
                        # --- TRY INSERT WITH config_data ---
                        supabase.table("crm_shared_history").insert(recs).execute()
                    except Exception as e:
                        # --- FALLBACK IF DB SCHEMA IS MISSING 'config_data' COLUMN ---
                        if "config_data" in str(e) or "PGRST204" in str(e):
                             # Remove 'config_data' key and retry insert
                             recs_fallback = [{k: v for k, v in r.items() if k != 'config_data'} for r in recs]
                             try:
                                 supabase.table("crm_shared_history").insert(recs_fallback).execute()
                                 st.warning("‚ö†Ô∏è ƒê√£ l∆∞u th√†nh c√¥ng (Ch·∫ø ƒë·ªô t∆∞∆°ng th√≠ch: B·ªè qua c·∫•u h√¨nh chi ph√≠ do Database c≈©).")
                             except Exception as e2:
                                 st.error(f"L·ªói Fatal sau khi retry: {e2}")
                                 st.stop()
                        else:
                             st.error(f"L·ªói l∆∞u Supabase: {e}")
                             st.stop()

                    # Save CSV Backup
                    try:
                        csv_buffer = io.BytesIO()
                        st.session_state.quote_df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
                        csv_buffer.seek(0)
                        csv_name = f"HIST_{quote_no}_{cust_name}_{int(time.time())}.csv"
                        curr_year = datetime.now().strftime("%Y")
                        curr_month = datetime.now().strftime("%b").upper()
                        path_list_hist = ["QUOTATION_HISTORY", cust_name, curr_year, curr_month]
                        lnk, _ = upload_to_drive_structured(csv_buffer, path_list_hist, csv_name)
                        
                        # --- NEW FEATURE: SAVE CONFIG FILE SEPARATELY TO DRIVE ---
                        # Creates an Excel file with the percentage configuration
                        df_cfg = pd.DataFrame([clean_params])
                        cfg_buffer = io.BytesIO()
                        df_cfg.to_excel(cfg_buffer, index=False)
                        cfg_buffer.seek(0)
                        cfg_name = f"CONFIG_{quote_no}_{cust_name}_{int(time.time())}.xlsx"
                        upload_to_drive_structured(cfg_buffer, path_list_hist, cfg_name)
                        
                        st.success("‚úÖ ƒê√£ l∆∞u l·ªãch s·ª≠ DB & CSV (K√®m file c·∫•u h√¨nh % ri√™ng)!")
                        st.markdown(f"üìÇ [Folder L·ªãch S·ª≠]({lnk})", unsafe_allow_html=True)
                    except Exception as e: st.error(f"L·ªói l∆∞u Drive: {e}")
                else: st.error("Ch·ªçn kh√°ch!")
            st.markdown('</div>', unsafe_allow_html=True)

# --- TAB 4: PO & ƒê·∫∂T H√ÄNG ---
with t4:
    if 'show_ncc_upload' not in st.session_state: st.session_state.show_ncc_upload = False
    if 'show_cust_upload' not in st.session_state: st.session_state.show_cust_upload = False
    if 'po_ncc_df' not in st.session_state: st.session_state.po_ncc_df = pd.DataFrame()
    if 'po_cust_df' not in st.session_state: st.session_state.po_cust_df = pd.DataFrame()
    
    st.markdown("### üîé TRA C·ª®U ƒê∆†N H√ÄNG (PO)")
    search_po = st.text_input("Nh·∫≠p s·ªë PO, M√£ h√†ng, T√™n h√†ng, Kh√°ch, NCC...", key="search_po_tab")
    if search_po:
        df_po_cust = load_data("db_customer_orders")
        df_po_supp = load_data("db_supplier_orders")
        res_cust = pd.DataFrame()
        if not df_po_cust.empty:
            mask_c = df_po_cust.astype(str).apply(lambda x: x.str.contains(search_po, case=False, na=False)).any(axis=1)
            res_cust = df_po_cust[mask_c]
            if not res_cust.empty:
                st.info(f"T√¨m th·∫•y {len(res_cust)} d√≤ng trong PO Kh√°ch H√†ng")
                st.dataframe(res_cust, use_container_width=True)
        res_supp = pd.DataFrame()
        if not df_po_supp.empty:
            mask_s = df_po_supp.astype(str).apply(lambda x: x.str.contains(search_po, case=False, na=False)).any(axis=1)
            res_supp = df_po_supp[mask_s]
            if not res_supp.empty:
                st.info(f"T√¨m th·∫•y {len(res_supp)} d√≤ng trong PO Nh√† Cung C·∫•p")
                st.dataframe(res_supp, use_container_width=True)
        if res_cust.empty and res_supp.empty: st.warning("Kh√¥ng t√¨m th·∫•y k·∫øt qu·∫£ n√†o.")

    st.divider()
    c_ncc, c_kh = st.columns(2)
    with c_ncc:
        st.subheader("1. ƒê·∫∂T H√ÄNG NH√Ä CUNG C·∫§P")
        with st.expander("üîê Admin: Reset ƒê·∫øm ƒê∆°n NCC"):
            adm_po_ncc = st.text_input("Pass Admin NCC", type="password")
            if st.button("Reset ƒê·∫øm ƒê∆°n NCC"):
                if adm_po_ncc == "admin":
                    supabase.table("db_supplier_orders").delete().neq("id", 0).execute()
                    st.success("ƒê√£ reset b·ªô ƒë·∫øm PO NCC!"); time.sleep(1); st.rerun()
                else: st.error("Sai Pass")

        if st.button("‚ûï T·∫†O M·ªöI (ƒê·∫∑t NCC)"):
            st.session_state.po_ncc_df = pd.DataFrame()
            st.session_state.show_ncc_upload = True 
            st.rerun()

        if st.session_state.show_ncc_upload:
            po_s_no = st.text_input("S·ªë PO NCC", key="po_s_input")
            up_s = st.file_uploader("Upload File Items (Excel)", key="ups")
            if up_s and st.button("Load Items NCC"):
                df_up = pd.read_excel(up_s, dtype=str).fillna("")
                db = load_data("crm_purchases")
                lookup = {clean_key(r['item_code']): r for r in db.to_dict('records')}
                recs = []
                for i, r in df_up.iterrows():
                    code_raw = safe_str(r.iloc[1])
                    qty_val = to_float(r.iloc[4])
                    no_val = safe_str(r.iloc[0]) 
                    match = lookup.get(clean_key(code_raw))
                    if match:
                        name = match['item_name']; specs = match['specs']; supplier = match['supplier_name']
                        buy_rmb = to_float(match['buying_price_rmb']); rate = to_float(match['exchange_rate'])
                        buy_vnd = to_float(match['buying_price_vnd']); leadtime = match['leadtime']
                    else:
                        name = safe_str(r.iloc[2]); specs = safe_str(r.iloc[3]); supplier = "Unknown"
                        buy_rmb = 0; rate = 0; buy_vnd = 0; leadtime = "0"
                    eta = calc_eta(datetime.now(), leadtime)
                    recs.append({
                        "No": no_val, "Item code": code_raw, "Item name": name, "Specs": specs, "Q'ty": qty_val,
                        "Buying price(RMB)": fmt_num(buy_rmb), "Total buying price(RMB)": fmt_num(buy_rmb * qty_val),
                        "Exchange rate": fmt_num(rate),
                        "Buying price(VND)": fmt_num(buy_vnd), "Total buying price(VND)": fmt_num(buy_vnd * qty_val),
                        "Supplier": supplier, "ETA": eta
                    })
                st.session_state.po_ncc_df = pd.DataFrame(recs)
            
            if not st.session_state.po_ncc_df.empty:
                st.dataframe(st.session_state.po_ncc_df, use_container_width=True, hide_index=True)
                if st.button("üíæ X√ÅC NH·∫¨N ƒê·∫∂T H√ÄNG NCC"):
                    if not po_s_no: st.error("Thi·∫øu s·ªë PO")
                    else:
                        grouped = st.session_state.po_ncc_df.groupby("Supplier")
                        created_files = []
                        for supp_name, group in grouped:
                            if not supp_name: supp_name = "Unknown"
                            db_recs = []
                            for r in group.to_dict('records'):
                                db_recs.append({
                                    "po_number": po_s_no, "supplier": supp_name, "order_date": datetime.now().strftime("%d/%m/%Y"),
                                    "item_code": r["Item code"], "item_name": r["Item name"], "specs": r["Specs"],
                                    "qty": to_float(r["Q'ty"]), "total_vnd": to_float(r["Total buying price(VND)"]),
                                    "eta": r["ETA"]
                                })
                            supabase.table("db_supplier_orders").insert(db_recs).execute()
                            track_rec = {
                                "po_no": f"{po_s_no}_{supp_name}", "partner": supp_name, "status": "Ordered", "order_type": "NCC",
                                "last_update": datetime.now().strftime("%d/%m/%Y"), 
                                "eta": group.iloc[0]["ETA"]
                            }
                            supabase.table("crm_tracking").insert([track_rec]).execute()
                            wb = Workbook(); ws = wb.active; ws.title = "PO"
                            headers = ["No", "Item code", "Item name", "Specs", "Q'ty", "Buying(RMB)", "Total(RMB)", "Rate", "Buying(VND)", "Total(VND)", "Supplier", "ETA"]
                            ws.append(headers)
                            for r in group.to_dict('records'):
                                ws.append([r["No"], r["Item code"], r["Item name"], r["Specs"], r["Q'ty"], 
                                           r["Buying price(RMB)"], r["Total buying price(RMB)"], r["Exchange rate"],
                                           r["Buying price(VND)"], r["Total buying price(VND)"], r["Supplier"], r["ETA"]])
                            out = io.BytesIO(); wb.save(out); out.seek(0)
                            curr_year = datetime.now().strftime("%Y")
                            curr_month = datetime.now().strftime("%b").upper()
                            file_name = f"PO_{po_s_no}_{supp_name}.xlsx"
                            path_list = ["PO_NCC", curr_year, supp_name, curr_month]
                            lnk, _ = upload_to_drive_structured(out, path_list, file_name)
                            created_files.append((file_name, lnk, out)) 
                        st.success(f"‚úÖ ƒê√£ t·∫°o {len(created_files)} PO cho c√°c NCC!")
                        for fname, lnk, buffer in created_files:
                            c_d1, c_d2 = st.columns([2,1])
                            c_d1.markdown(f"üìÇ **[M·ªü Folder: {fname}]({lnk})**", unsafe_allow_html=True)
                            c_d2.download_button(label=f"üì• T·∫£i {fname}", data=buffer, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{fname}")

    with c_kh:
        st.subheader("2. PO KH√ÅCH H√ÄNG")
        with st.expander("üîê Admin: Reset ƒê·∫øm ƒê∆°n Kh√°ch"):
            adm_po_cust = st.text_input("Pass Admin Cust", type="password")
            if st.button("Reset ƒê·∫øm ƒê∆°n Kh√°ch"):
                if adm_po_cust == "admin":
                    supabase.table("db_customer_orders").delete().neq("id", 0).execute()
                    st.success("ƒê√£ reset b·ªô ƒë·∫øm PO Kh√°ch!"); time.sleep(1); st.rerun()
                else: st.error("Sai Pass")

        if st.button("‚ûï T·∫†O M·ªöI (PO Kh√°ch)"):
            st.session_state.po_cust_df = pd.DataFrame()
            st.session_state.show_cust_upload = True
            st.rerun()

        if st.session_state.show_cust_upload:
            po_c_no = st.text_input("S·ªë PO Kh√°ch", key="po_c_input")
            custs = load_data("crm_customers")
            c_name = st.selectbox("Ch·ªçn Kh√°ch", [""] + custs['short_name'].tolist() if not custs.empty else [], key="sel_cust_po")
            uploaded_files = st.file_uploader("Upload File PO Kh√°ch (Excel/PDF)", type=['xlsx', 'pdf'], accept_multiple_files=True, key="upc")
            if uploaded_files and st.button("Load PO Kh√°ch"):
                if not c_name: st.error("Vui l√≤ng ch·ªçn kh√°ch tr∆∞·ªõc ƒë·ªÉ l·∫•y gi√°!")
                else:
                    excel_files = [f for f in uploaded_files if f.name.endswith('.xlsx')]
                    if excel_files:
                        all_recs = []
                        hist = load_data("crm_shared_history") 
                        cust_hist = hist[hist['customer'] == c_name].sort_values(by='date', ascending=False)
                        price_lookup = {}
                        for _, h in cust_hist.iterrows():
                            c_code = clean_key(h['item_code'])
                            if c_code not in price_lookup: price_lookup[c_code] = to_float(h['unit_price'])
                        db_items = load_data("crm_purchases")
                        lt_lookup = {clean_key(r['item_code']): r['leadtime'] for r in db_items.to_dict('records')}
                        for f in excel_files:
                            try:
                                df_up = pd.read_excel(f, header=None, skiprows=1, dtype=str).fillna("")
                                for i, r in df_up.iterrows():
                                    no_val = safe_str(r.iloc[0]) 
                                    code = safe_str(r.iloc[1])
                                    qty = to_float(r.iloc[4])
                                    unit_price = price_lookup.get(clean_key(code), 0)
                                    total = unit_price * qty
                                    leadtime = lt_lookup.get(clean_key(code), "0")
                                    eta = calc_eta(datetime.now(), leadtime)
                                    if code:
                                         all_recs.append({
                                             "No.": no_val, "Item code": code, "Item name": safe_str(r.iloc[2]),
                                             "Specs": safe_str(r.iloc[3]), "Q'ty": qty,
                                             "Unit price(VND)": fmt_num(unit_price), "Total price(VND)": fmt_num(total),
                                             "Customer": c_name, "ETA": eta, "Source File": f.name
                                         })
                            except: pass
                        st.session_state.po_cust_df = pd.DataFrame(all_recs)
                    else: st.info("Ch·ªâ load data t·ª´ Excel. PDF s·∫Ω ƒë∆∞·ª£c l∆∞u khi b·∫•m 'L∆∞u PO'.")

            if not st.session_state.po_cust_df.empty:
                st.dataframe(st.session_state.po_cust_df, use_container_width=True, hide_index=True)
                if st.button("üíæ L∆ØU PO KH√ÅCH H√ÄNG"):
                    if not po_c_no: st.error("Thi·∫øu s·ªë PO")
                    else:
                        db_recs = []
                        for r in st.session_state.po_cust_df.to_dict('records'):
                            db_recs.append({
                                "po_number": po_c_no, "customer": c_name, "order_date": datetime.now().strftime("%d/%m/%Y"),
                                "item_code": r["Item code"], "item_name": r["Item name"], "specs": r["Specs"],
                                "qty": to_float(r["Q'ty"]), "unit_price": to_float(r["Unit price(VND)"]),
                                "total_price": to_float(r["Total price(VND)"]), "eta": r["ETA"]
                            })
                        supabase.table("db_customer_orders").insert(db_recs).execute()
                        track_rec = {
                            "po_no": po_c_no, "partner": c_name, "status": "Waiting", "order_type": "KH",
                            "last_update": datetime.now().strftime("%d/%m/%Y"),
                            "eta": st.session_state.po_cust_df.iloc[0]["ETA"]
                        }
                        supabase.table("crm_tracking").insert([track_rec]).execute()
                        curr_year = datetime.now().strftime("%Y")
                        curr_month = datetime.now().strftime("%b").upper()
                        path_list = ["PO_KHACH_HANG", curr_year, c_name, curr_month]
                        saved_links = []
                        if uploaded_files:
                            for upf in uploaded_files:
                                upf.seek(0)
                                f_name = f"{po_c_no}_{upf.name}"
                                lnk, _ = upload_to_drive_structured(upf, path_list, f_name)
                                saved_links.append(lnk)
                        st.success("‚úÖ L∆∞u PO Kh√°ch th√†nh c√¥ng! ƒê√£ link sang Tracking.")
                        if saved_links:
                             st.markdown(f"üìÇ **[M·ªü Folder PO Kh√°ch: {c_name}/{curr_month}]({saved_links[0]})**", unsafe_allow_html=True)

# --- TAB 5: TRACKING ---
with t5:
    st.subheader("THEO D√ïI ƒê∆†N H√ÄNG (TRACKING)")
    if st.button("üîÑ Refresh Tracking"): st.cache_data.clear(); st.rerun()
    df_track = load_data("crm_tracking", order_by="id")
    if not df_track.empty:
        c1, c2 = st.columns([1, 2])
        with c1:
            st.markdown("#### C·∫≠p nh·∫≠t tr·∫°ng th√°i / ·∫¢nh")
            po_list = df_track['po_no'].unique()
            sel_po = st.selectbox("Ch·ªçn PO", po_list, key="tr_po")
            new_status = st.selectbox("Tr·∫°ng th√°i m·ªõi", ["Ordered", "Shipping", "Arrived", "Delivered", "Waiting"], key="tr_st")
            proof_img = st.file_uploader("Upload ·∫¢nh Proof", type=['png', 'jpg'], key="tr_img")
            if st.button("C·∫≠p nh·∫≠t Tracking"):
                upd_data = {"status": new_status, "last_update": datetime.now().strftime("%d/%m/%Y")}
                if proof_img:
                    lnk, _ = upload_to_drive_simple(proof_img, "CRM_PROOF", f"PRF_{sel_po}_{int(time.time())}.png")
                    upd_data["proof_image"] = lnk
                supabase.table("crm_tracking").update(upd_data).eq("po_no", sel_po).execute()
                st.success("Updated!"); time.sleep(1); st.rerun()
        with c2:
            st.markdown("#### Danh s√°ch ƒë∆°n h√†ng")
            st.dataframe(
                df_track, 
                column_config={
                    "proof_image": st.column_config.ImageColumn("Proof"), 
                    "status": st.column_config.TextColumn("Status"),
                    "po_no": "PO No.", "partner": "Partner", "eta": "ETA"
                }, 
                use_container_width=True, hide_index=True
            )
    else: st.info("Ch∆∞a c√≥ d·ªØ li·ªáu Tracking. H√£y t·∫°o PO ·ªü Tab 4.")

# --- TAB 6: MASTER DATA ---
with t6:
    tc, ts, tt = st.tabs(["KH√ÅCH H√ÄNG", "NH√Ä CUNG C·∫§P", "TEMPLATE"])
    
    # --- FIX: DYNAMIC CUSTOMER IMPORT ---
    with tc:
        df_c = load_data("crm_customers")
        st.write("D·ªØ li·ªáu hi·ªán t·∫°i:")
        st.dataframe(df_c, use_container_width=True, hide_index=True)
        
        up = st.file_uploader("Import CUSTOMER LIST", type=["xlsx"], key="uck")
        if up and st.button("üöÄ Import KH (ƒê·ªìng b·ªô tuy·ªát ƒë·ªëi)"):
            try:
                # ƒê·ªçc Excel (L·∫•y lu√¥n header t·ª´ file)
                df_new = pd.read_excel(up, dtype=str).fillna("")
                
                # Clean t√™n c·ªôt (x√≥a kho·∫£ng tr·∫Øng th·ª´a)
                df_new.columns = [str(c).strip() for c in df_new.columns]
                
                # Chuy·ªÉn th√†nh list dict ƒë·ªÉ insert
                records = df_new.to_dict('records')
                
                if records:
                    # 1. X√ìA S·∫†CH D·ªÆ LI·ªÜU C≈® (ƒê·ªÉ ƒë·ªìng b·ªô tuy·ªát ƒë·ªëi nh∆∞ y√™u c·∫ßu)
                    supabase.table("crm_customers").delete().neq("id", 0).execute()
                    
                    # 2. INSERT D·ªÆ LI·ªÜU M·ªöI (DYNAMIC COLUMNS)
                    # L∆∞u √Ω: Database Supabase PH·∫¢I c√≥ c√°c c·ªôt t∆∞∆°ng ·ª©ng v·ªõi header trong Excel
                    chunk_size = 100
                    for k in range(0, len(records), chunk_size):
                        batch = records[k:k+chunk_size]
                        supabase.table("crm_customers").insert(batch).execute()
                        
                    st.success(f"‚úÖ ƒê√£ ƒë·ªìng b·ªô {len(records)} kh√°ch h√†ng! (C·∫•u tr√∫c c·ªôt theo Excel)")
                    time.sleep(1); st.rerun()
                else: st.warning("File r·ªóng!")
            except Exception as e:
                st.error(f"L·ªói Import: {e}")
                st.warning("‚ö†Ô∏è L∆∞u √Ω: T√™n c·ªôt trong file Excel PH·∫¢I kh·ªõp v·ªõi t√™n c·ªôt trong Database Supabase.")

    # --- FIX: DYNAMIC SUPPLIER IMPORT ---
    with ts:
        df_s = load_data("crm_suppliers")
        st.write("D·ªØ li·ªáu hi·ªán t·∫°i:")
        st.dataframe(df_s, use_container_width=True, hide_index=True)
        
        up_s = st.file_uploader("Import SUPPLIER LIST", type=["xlsx"], key="usn")
        if up_s and st.button("üöÄ Import NCC (ƒê·ªìng b·ªô tuy·ªát ƒë·ªëi)"):
            try:
                # ƒê·ªçc Excel
                df_new = pd.read_excel(up_s, dtype=str).fillna("")
                
                # Clean t√™n c·ªôt
                df_new.columns = [str(c).strip() for c in df_new.columns]
                
                records = df_new.to_dict('records')
                
                if records:
                    # 1. X√ìA S·∫†CH D·ªÆ LI·ªÜU C≈®
                    supabase.table("crm_suppliers").delete().neq("id", 0).execute()
                    
                    # 2. INSERT D·ªÆ LI·ªÜU M·ªöI
                    chunk_size = 100
                    for k in range(0, len(records), chunk_size):
                        batch = records[k:k+chunk_size]
                        supabase.table("crm_suppliers").insert(batch).execute()
                        
                    st.success(f"‚úÖ ƒê√£ ƒë·ªìng b·ªô {len(records)} nh√† cung c·∫•p! (C·∫•u tr√∫c c·ªôt theo Excel)")
                    time.sleep(1); st.rerun()
                else: st.warning("File r·ªóng!")
            except Exception as e:
                st.error(f"L·ªói Import: {e}")
                st.warning("‚ö†Ô∏è L∆∞u √Ω: T√™n c·ªôt trong file Excel PH·∫¢I kh·ªõp v·ªõi t√™n c·ªôt trong Database Supabase.")

    with tt:
        st.write("Upload Template Excel")
        up_t = st.file_uploader("File Template (.xlsx)", type=["xlsx"])
        t_name = st.text_input("T√™n Template (Nh·∫≠p: AAA-QUOTATION)")
        if up_t and t_name and st.button("L∆∞u Template"):
            lnk, fid = upload_to_drive_simple(up_t, "CRM_TEMPLATES", f"TMP_{t_name}.xlsx")
            if fid: supabase.table("crm_templates").insert([{"template_name": t_name, "file_id": fid, "last_updated": datetime.now().strftime("%d/%m/%Y")}]).execute(); st.success("OK"); st.rerun()
        st.dataframe(load_data("crm_templates"))
